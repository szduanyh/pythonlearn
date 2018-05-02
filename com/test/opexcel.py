#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
http://openpyxl.readthedocs.io/en/default/usage.html
"""

# workbook相关
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = '测试.xlsx'

ws1 = wb.active
ws1.title = "测试1"

for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row,
                     value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
wb.save(filename=dest_filename)
